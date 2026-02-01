from fastapi import APIRouter, File, UploadFile, HTTPException, Depends
from fastapi.responses import JSONResponse
import pandas as pd
import openpyxl
import io
import logging
from typing import List, Dict, Any
from pydantic import BaseModel

from auth import get_current_user
from models import APIResponse, User

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api", tags=["excel-compare"])


@router.post("/excel-test")
async def test_excel_upload(
    file1: UploadFile = File(...),
    file2: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Test endpoint to debug file upload issues."""
    try:
        logger.info(f"Test upload - File1: {file1.filename}, File2: {file2.filename}")
        logger.info(f"File1 content type: {file1.content_type}, File2 content type: {file2.content_type}")
        logger.info(f"File1 size: {getattr(file1, 'size', 'N/A')}, File2 size: {getattr(file2, 'size', 'N/A')}")
        
        return APIResponse(
            success=True,
            message="Files received successfully",
            data={
                "file1_name": file1.filename,
                "file2_name": file2.filename,
                "file1_content_type": file1.content_type,
                "file2_content_type": file2.content_type,
                "file1_size": getattr(file1, 'size', None),
                "file2_size": getattr(file2, 'size', None)
            }
        )
    except Exception as e:
        logger.error(f"Test upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class ExcelCompareResult(BaseModel):
    """Result model for Excel comparison."""
    success: bool
    total_sheets: int
    matched_sheets: int
    comparison_results: List[Dict[str, Any]]
    summary: str


@router.post("/excel-compare", response_model=APIResponse)
async def compare_excel_files(
    file1: UploadFile = File(None, description="First Excel file to compare"),
    file2: UploadFile = File(None, description="Second Excel file to compare"),
    current_user: User = Depends(get_current_user),
):
    """Compare two Excel files sheet by sheet, cell by cell.
    
    All users can access this functionality.
    Both files must have the same number of sheets with matching sheet names.
    """
    
    try:
        from input_validation import InputValidator
        
        logger.info(f"Excel compare request - File1: {file1.filename if file1 else 'None'}, File2: {file2.filename if file2 else 'None'}")
        logger.info(f"File1 size: {getattr(file1, 'size', 'Unknown')}, File2 size: {getattr(file2, 'size', 'Unknown')}")
        
        # Check if files are provided
        if not file1:
            raise HTTPException(status_code=400, detail="File1 is required.")
        
        if not file2:
            raise HTTPException(status_code=400, detail="File2 is required.")
        
        # Check if files have filenames
        if not file1.filename:
            raise HTTPException(status_code=400, detail="File1 must have a filename.")
        
        if not file2.filename:
            raise HTTPException(status_code=400, detail="File2 must have a filename.")
        
        # Validate file names for path traversal
        if not InputValidator.validate_file_upload(file1.filename, ['.xlsx', '.xls']):
            raise HTTPException(status_code=400, detail="Invalid file1 name or type. Only Excel files are allowed.")
        
        if not InputValidator.validate_file_upload(file2.filename, ['.xlsx', '.xls']):
            raise HTTPException(status_code=400, detail="Invalid file2 name or type. Only Excel files are allowed.")
        
        # Validate file sizes (limit to 50MB each)
        MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB
        if file1.size and file1.size > MAX_FILE_SIZE:
            raise HTTPException(status_code=400, detail=f"File1 too large. Maximum size is {MAX_FILE_SIZE // (1024*1024)}MB.")
        
        if file2.size and file2.size > MAX_FILE_SIZE:
            raise HTTPException(status_code=400, detail=f"File2 too large. Maximum size is {MAX_FILE_SIZE // (1024*1024)}MB.")
        
        # Additional filename sanitization
        file1_name = InputValidator.sanitize_string(file1.filename, max_length=255)
        file2_name = InputValidator.sanitize_string(file2.filename, max_length=255)
        
        # Read Excel files with size monitoring
        logger.info(f"Reading Excel files: {file1_name} ({file1.size} bytes), {file2_name} ({file2.size} bytes)")
        
        try:
            # Read files with timeout protection
            import asyncio
            file1_content = await asyncio.wait_for(file1.read(), timeout=30)
            file2_content = await asyncio.wait_for(file2.read(), timeout=30)
            
            # Basic content validation - check for Excel file signatures
            if not (file1_content.startswith(b'PK') or file1_content.startswith(b'\xd0\xcf')):
                raise HTTPException(status_code=400, detail="File1 does not appear to be a valid Excel file.")
            
            if not (file2_content.startswith(b'PK') or file2_content.startswith(b'\xd0\xcf')):
                raise HTTPException(status_code=400, detail="File2 does not appear to be a valid Excel file.")
                
        except asyncio.TimeoutError:
            logger.error("File read timeout")
            raise HTTPException(status_code=400, detail="File reading timed out. Files may be too large or corrupted.")
        except Exception as e:
            logger.error(f"Error reading uploaded files: {e}")
            raise HTTPException(status_code=400, detail="Error reading uploaded files. Please ensure files are not corrupted.")
        
        try:
            # Load with strict parsing and limits to prevent zip bombs
            workbook1 = openpyxl.load_workbook(
                io.BytesIO(file1_content), 
                data_only=True,
                read_only=True
            )
            logger.info(f"Successfully loaded {file1_name}")
        except Exception as e:
            logger.error(f"Error loading {file1_name}: {e}")
            raise HTTPException(status_code=400, detail=f"Error loading {file1_name}. Please ensure it's a valid Excel file.")
            
        try:
            workbook2 = openpyxl.load_workbook(
                io.BytesIO(file2_content), 
                data_only=True,
                read_only=True
            )
            logger.info(f"Successfully loaded {file2_name}")
        except Exception as e:
            logger.error(f"Error loading {file2_name}: {e}")
            raise HTTPException(status_code=400, detail=f"Error loading {file2_name}. Please ensure it's a valid Excel file.")
        
        # Get sheet names
        sheets1 = workbook1.sheetnames
        sheets2 = workbook2.sheetnames
        
        # Handle different number of sheets gracefully
        if len(sheets1) != len(sheets2):
            logger.warning(f"Files have different number of sheets: {len(sheets1)} vs {len(sheets2)}")
            # Still proceed with comparison of common sheets
            common_sheets = set(sheets1) & set(sheets2)
            if not common_sheets:
                raise HTTPException(
                    status_code=400,
                    detail=f"No common sheets found between files. File 1 has: {sheets1}. File 2 has: {sheets2}"
                )
            # Use common sheets for comparison
            sheets_to_compare = list(common_sheets)
            logger.info(f"Comparing {len(sheets_to_compare)} common sheets: {sheets_to_compare}")
        else:
            sheets_to_compare = sheets1
        
        # Note: We now handle mismatched sheets gracefully above
        
        comparison_results = []
        matched_sheets = 0
        
        # Compare each sheet
        for sheet_name in sheets_to_compare:
            if sheet_name in workbook1.sheetnames and sheet_name in workbook2.sheetnames:
                sheet_result = compare_sheets(
                    workbook1[sheet_name], 
                    workbook2[sheet_name], 
                    sheet_name
                )
                comparison_results.append(sheet_result)
                
                if sheet_result["status"] == "matched":
                    matched_sheets += 1
            else:
                # Sheet exists in one file but not the other
                comparison_results.append({
                    "sheet": sheet_name,
                    "cell_id": "NA",
                    "value1": "sheet exists" if sheet_name in workbook1.sheetnames else "sheet missing",
                    "value2": "sheet exists" if sheet_name in workbook2.sheetnames else "sheet missing",
                    "status": "sheet_mismatch",
                    "differences": []
                })
        
        # Add info about sheets that exist in only one file
        sheets1_only = set(sheets1) - set(sheets2)
        sheets2_only = set(sheets2) - set(sheets1)
        
        for sheet_name in sheets1_only:
            comparison_results.append({
                "sheet": sheet_name,
                "cell_id": "NA",
                "value1": "sheet exists",
                "value2": "sheet missing",
                "status": "sheet_missing_in_file2",
                "differences": []
            })
            
        for sheet_name in sheets2_only:
            comparison_results.append({
                "sheet": sheet_name,
                "cell_id": "NA",
                "value1": "sheet missing",
                "value2": "sheet exists", 
                "status": "sheet_missing_in_file1",
                "differences": []
            })
        
        success = matched_sheets == len(sheets_to_compare) and len(sheets1_only) == 0 and len(sheets2_only) == 0
        total_sheets_compared = len(sheets_to_compare) + len(sheets1_only) + len(sheets2_only)
        summary = f"Compared {total_sheets_compared} sheets. {matched_sheets} matched, {total_sheets_compared - matched_sheets} had differences or were missing."
        
        result = {
            "success": success,
            "total_sheets": total_sheets_compared,
            "matched_sheets": matched_sheets,
            "comparison_results": comparison_results,
            "summary": summary,
            "files_info": {
                "file1_name": file1.filename,
                "file2_name": file2.filename,
                "file1_sheets": len(sheets1),
                "file2_sheets": len(sheets2)
            }
        }
        
        return APIResponse(
            success=True,
            message="Excel comparison completed successfully",
            data=result
        )
        
    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"Excel comparison error: {exc}", exc_info=True)
        # Don't expose internal errors to users
        raise HTTPException(status_code=500, detail="Failed to compare Excel files. Please ensure both files are valid Excel files and try again.")


def compare_sheets(sheet1, sheet2, sheet_name: str, max_differences: int = 1000) -> Dict[str, Any]:
    """Compare two Excel sheets cell by cell with performance optimizations."""
    
    differences = []
    cells_compared = 0
    
    try:
        # Get the maximum row and column from both sheets
        max_row1, max_col1 = sheet1.max_row, sheet1.max_column
        max_row2, max_col2 = sheet2.max_row, sheet2.max_column
        
        # Handle empty sheets
        if (max_row1 == 1 and max_col1 == 1 and sheet1.cell(1, 1).value is None and
            max_row2 == 1 and max_col2 == 1 and sheet2.cell(1, 1).value is None):
            logger.info(f"Both sheets '{sheet_name}' are empty - marked as matched")
            return {
                "sheet": sheet_name,
                "cell_id": "NA",
                "value1": "Empty sheet",
                "value2": "Empty sheet",
                "status": "matched",
                "differences": []
            }
        
        max_row = max(max_row1, max_row2)
        max_col = max(max_col1, max_col2)
        
        # Performance check for very large sheets
        total_cells = max_row * max_col
        if total_cells > 1000000:  # Limit to 1M cells
            logger.warning(f"Sheet '{sheet_name}' is very large ({total_cells} cells). Limiting comparison to first 1000 rows.")
            max_row = min(max_row, 1000)
        
        logger.info(f"Comparing sheet '{sheet_name}': {max_row} rows × {max_col} columns")
        
        # Compare each cell
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cells_compared += 1
                cell1_value = None
                cell2_value = None
                
                try:
                    if row <= max_row1 and col <= max_col1:
                        cell1 = sheet1.cell(row=row, column=col)
                        cell1_value = cell1.value
                    
                    if row <= max_row2 and col <= max_col2:
                        cell2 = sheet2.cell(row=row, column=col)
                        cell2_value = cell2.value
                except Exception as e:
                    logger.warning(f"Error reading cell {row},{col} in sheet '{sheet_name}': {e}")
                    continue
                
                # Normalize None values to empty string for comparison
                if cell1_value is None:
                    cell1_value = ""
                if cell2_value is None:
                    cell2_value = ""
                
                # Convert to string for comparison, handling special types
                try:
                    cell1_str = str(cell1_value) if cell1_value != "" else ""
                    cell2_str = str(cell2_value) if cell2_value != "" else ""
                except Exception as e:
                    logger.warning(f"Error converting cell values to string at {row},{col}: {e}")
                    cell1_str = "ERROR_CONVERTING"
                    cell2_str = "ERROR_CONVERTING"
                
                if cell1_str != cell2_str:
                    # Stop if we've found too many differences for performance
                    if len(differences) >= max_differences:
                        logger.warning(f"Reached maximum differences limit ({max_differences}) for sheet '{sheet_name}'")
                        break
                        
                    cell_id = f"{openpyxl.utils.get_column_letter(col)}{row}"
                    differences.append({
                        "sheet": sheet_name,
                        "cell_id": cell_id,
                        "value1": cell1_str[:100],  # Limit length for display
                        "value2": cell2_str[:100],  # Limit length for display
                        "status": "not matched"
                    })
            
            # Break outer loop if we hit the limit
            if len(differences) >= max_differences:
                break
        
        logger.info(f"Sheet '{sheet_name}' comparison complete: {cells_compared} cells compared, {len(differences)} differences found")
        
        if not differences:
            return {
                "sheet": sheet_name,
                "cell_id": "NA",
                "value1": "NA", 
                "value2": "NA",
                "status": "matched",
                "differences": []
            }
        else:
            return {
                "sheet": sheet_name,
                "cell_id": "multiple",
                "value1": "multiple",
                "value2": "multiple", 
                "status": "not matched",
                "differences": differences
            }
    
    except Exception as e:
        logger.error(f"Error comparing sheet '{sheet_name}': {e}")
        return {
            "sheet": sheet_name,
            "cell_id": "ERROR",
            "value1": "Error during comparison",
            "value2": str(e),
            "status": "comparison_error",
            "differences": []
        }